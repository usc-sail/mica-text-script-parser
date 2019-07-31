import os
import pickle
from openpyxl import load_workbook
from ._read_annotation_movies import read_annotation_movies
from ._check_empty_cell import is_cell_empty

LABELS = ["S","N","C","D","E","T","M","!"," "]

def read_annotations(annotation_filepath):
    print("reading from {}...".format(annotation_filepath))

    pickle_filepath = annotation_filepath.replace(".xlsx", ".pkl")
    
    if os.path.exists(pickle_filepath):
        annotations = pickle.load(open(pickle_filepath, "rb"))
    else:
        wb = load_workbook(annotation_filepath)
        _, begs, ends = read_annotation_movies()
        annotations = []

        for sheetname, beg, end in zip(wb.sheetnames[1:], begs, ends):
            ws = wb[sheetname]
            labels = ""
            for row in range(3, end - beg + 4):
                if not is_cell_empty(ws, row, 1):
                    column = 3
                    while column < 10:
                        if not is_cell_empty(ws, row, column):
                            break
                        column += 1
                    labels += LABELS[column - 3]
                else:
                    labels += " "
            annotations.append(labels)
        pickle.dump(annotations, open(pickle_filepath, "wb"))

    return annotations

if __name__ == "__main__":
    annotations = read_annotations("../data/annotations/Annotator 2.xlsx")
    print(len(annotations))