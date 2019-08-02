# author - Sabyasachee

from openpyxl import load_workbook

def is_cell_empty(ws, i, j):
    # return True if cell i, j in worksheet ws is empty, else False
    value = ws.cell(i, j).value
    return value is None or str(value).strip() == ""

def check_annotation_ranges():
    # check whether the text between the beginning and ending ranges in data/annotation_movies.txt in the corresponding
    # movie script is same as the content in the annotator xlsx sheet
    # here we check only for Annotator 1 (Ming), just replace Annotator 1 with Annotator 2 for Saby, or Annotator 3 for Tian
    # 
    # prints the total number of lines and the 10 starting and ending lines for both script and xlsx for each movie
    # user is required to manually check the output to ensure that everything is correct

    with open("data/annotation_movies.txt") as fr:
        lines = fr.read().split("\n")
        movies = []; begs = []; ends = []
        for line in lines:
            movie, beg, end = line.split()
            movies.append(movie.strip())
            begs.append(int(beg.strip()))
            ends.append(int(end.strip()))

    wb = load_workbook("data/annotations/Annotator 1.xlsx")
    sheetnames = wb.sheetnames[1:]

    for movie, beg, end, sheetname in zip(movies, begs, ends, sheetnames):
        with open("../../mica-scripts/scripts_txt/{}.txt".format(movie)) as fr:
            script = fr.read().split("\n")
            script_text = "\n".join(script[beg - 1: beg + 9])
            script_text += "..\n..\n..\n" + "\n".join(script[end - 10: end])
        
        ws = wb[sheetname]
        n_rows = ws.max_row
        while is_cell_empty(ws, n_rows, 1):
            n_rows -= 1

        anno_text = "\n".join(["" if is_cell_empty(ws, 3 + i, 1) else str(ws.cell(3 + i, 1).value) for i in range(10)])
        anno_text += "..\n..\n..\n" + "\n".join(["" if is_cell_empty(ws, n_rows - i, 1) else str(ws.cell(n_rows - i, 1).value) for i in range(9, -1, -1)])

        print("{}: SCRIPT = {}, ANNOTATION = {}".format(movie, end - beg + 1, n_rows - 2))
        print("\n\nSCRIPT => \n\n************************\n{}\n************************".format(script_text))
        print("\n\nANNOTATION => \n\n************************\n{}\n************************".format(anno_text))
        print("\n\n||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||\n\n")
        
if __name__ == "__main__":
    check_annotation_ranges()